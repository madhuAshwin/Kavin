"""
PM Deposition Modelling in Human Respiratory System
====================================================
Monte Carlo simulation of particulate matter (PM 0.1 - PM 10) deposition
in the human lung, based on the stochastic IDEAL-2 model (Koblinger & Hofmann).

Reference paper: "A Study of Environmental Impact of Air Pollution on Human Health:
PM Deposition Modelling" - Mayilvahanan I, Naresh K & Gandhimathi A (KCT, Coimbatore)

Author of this implementation: Claude (reconstructed from paper methodology)
"""

import argparse
import copy
import queue
import threading
import traceback
from dataclasses import dataclass
from pathlib import Path as FilePath
from tkinter import filedialog, messagebox, scrolledtext, ttk
import tkinter as tk
from typing import Callable, Dict, List, Optional, Tuple

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch, Ellipse, Arc, PathPatch, FancyArrowPatch
from matplotlib.path import Path
from mpl_toolkits.mplot3d import Axes3D
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk

# ============================================================================
# CONSTANTS
# ============================================================================

BOLTZMANN_K = 1.380649e-23       # J/K
TEMPERATURE = 310.15             # K (37°C body temperature)
AIR_VISCOSITY = 1.81e-5          # Pa·s (dynamic viscosity of air at 37°C)
AIR_DENSITY = 1.125              # kg/m³ (air density at 37°C)
PARTICLE_DENSITY = 1000.0        # kg/m³ (unit density as per paper: 1.0 g/cm³)
GRAVITY = 9.81                   # m/s²
MEAN_FREE_PATH = 0.066e-6        # m (mean free path of air molecules)

DEFAULT_PM_SIZES = [0.1, 0.5, 1.0, 1.5, 2.0, 2.5, 10.0]
DEFAULT_N_PARTICLES = 10000
DEFAULT_SEED = 42
DEFAULT_AGE_GROUPS = {
    'D (21-30 yrs)': {'tidal_volume': 2800e-6, 'breathing_freq': 15, 'label': '21-30'},
    'C (31-35 yrs)': {'tidal_volume': 2600e-6, 'breathing_freq': 14, 'label': '31-35'},
    'B (36-40 yrs)': {'tidal_volume': 2400e-6, 'breathing_freq': 13, 'label': '36-40'},
    'A (41-50 yrs)': {'tidal_volume': 2200e-6, 'breathing_freq': 12, 'label': '41-50'},
}

DEFAULT_TEMPLATE_NAME = 'pollution_input_template.xlsx'
OUTPUT_FILENAMES = {
    'bars': 'pm_deposition_bars.png',
    'summary': 'pm_results_summary.txt',
}
AGE_GROUP_OUTPUT_PATTERNS = {
    'scatter': 'pm_3d_scatter_{slug}.png',
    'heatmap': 'pm_heatmap_{slug}.png',
    'infographic': 'pm_lung_infographic_{slug}.png',
}
WORKBOOK_SHEETS = {
    'settings': 'simulation_settings',
    'pm_sizes': 'pm_sizes',
    'age_groups': 'age_groups',
}


@dataclass
class AgeGroupConfig:
    """Human-readable simulation input for a breathing cohort."""
    key: str
    tidal_volume_cm3: float
    breathing_freq: float
    label: str

    @property
    def tidal_volume_m3(self) -> float:
        return self.tidal_volume_cm3 * 1e-6

    def as_runtime_dict(self) -> Dict[str, float | str]:
        return {
            'tidal_volume': self.tidal_volume_m3,
            'breathing_freq': self.breathing_freq,
            'label': self.label,
        }


@dataclass
class SimulationConfig:
    """Runtime configuration shared by CLI, workbook loading, and the UI."""
    n_particles: int
    pm_sizes: List[float]
    age_groups: Dict[str, AgeGroupConfig]
    seed: int = DEFAULT_SEED

    def runtime_age_groups(self) -> Dict[str, Dict[str, float | str]]:
        return {
            age_key: age_group.as_runtime_dict()
            for age_key, age_group in self.age_groups.items()
        }


def get_default_config() -> SimulationConfig:
    """Return the default configuration that mirrors the original script."""
    age_groups = {
        age_key: AgeGroupConfig(
            key=age_key,
            tidal_volume_cm3=age_data['tidal_volume'] * 1e6,
            breathing_freq=age_data['breathing_freq'],
            label=age_data['label'],
        )
        for age_key, age_data in DEFAULT_AGE_GROUPS.items()
    }
    return SimulationConfig(
        n_particles=DEFAULT_N_PARTICLES,
        pm_sizes=list(DEFAULT_PM_SIZES),
        age_groups=age_groups,
        seed=DEFAULT_SEED,
    )


def clone_config(config: SimulationConfig) -> SimulationConfig:
    """Return a copy that can be mutated without affecting the defaults."""
    return copy.deepcopy(config)


def emit_log(message: str, log_callback: Optional[Callable[[str], None]] = None):
    """Send a log line to either the UI callback or stdout."""
    if log_callback is not None:
        log_callback(message)
    else:
        try:
            print(message)
        except UnicodeEncodeError:
            print(message.encode('ascii', errors='replace').decode('ascii'))


def validate_numeric(value, field_name: str, *, allow_zero: bool = False) -> float:
    """Parse numeric workbook cells with a clear validation error."""
    try:
        numeric_value = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f'{field_name} must be numeric, got {value!r}.') from exc

    if allow_zero:
        if numeric_value < 0:
            raise ValueError(f'{field_name} must be zero or greater, got {numeric_value}.')
    elif numeric_value <= 0:
        raise ValueError(f'{field_name} must be greater than zero, got {numeric_value}.')
    return numeric_value


def normalize_age_label(raw_label: str, fallback_key: str) -> str:
    """Use a readable label in plots even if the workbook leaves it blank."""
    text = (raw_label or '').strip()
    return text if text else fallback_key


def get_primary_age_key(config: SimulationConfig) -> str:
    """Prefer the original paper's D-category when it exists."""
    if 'D (21-30 yrs)' in config.age_groups:
        return 'D (21-30 yrs)'
    return next(iter(config.age_groups))

# Lung regions mapped to airway generations
REGION_MAP = {
    'Extra-thoracic': (0, 3),       # Nose, pharynx, larynx
    'Conducting (Tubular)': (3, 16), # Trachea through terminal bronchioles
    'Alveolar': (16, 23),            # Respiratory bronchioles through alveolar sacs
}


# ============================================================================
# WEIBEL SYMMETRIC LUNG MODEL (Model A, 23 generations)
# ============================================================================

@dataclass
class AirwayGeneration:
    """Properties of a single airway generation."""
    generation: int
    length: float        # meters
    diameter: float      # meters
    n_airways: int       # number of parallel airways
    branching_angle: float  # radians

def build_weibel_lung() -> List[AirwayGeneration]:
    """
    Build Weibel Type A symmetric lung model.
    Data from Weibel (1963) "Morphometry of the Human Lung".
    """
    # Weibel model data: (generation, length_mm, diameter_mm, branching_angle_deg)
    weibel_data = [
        (0,  120.0,  18.0,   0),    # Trachea
        (1,   47.6,  12.2,  35),    # Main bronchi
        (2,   19.0,   8.3,  30),    # Lobar bronchi
        (3,    7.6,   5.6,  25),    # Segmental bronchi
        (4,   12.7,   4.5,  25),
        (5,   10.7,   3.5,  22),
        (6,    9.0,   2.8,  20),
        (7,    7.6,   2.3,  18),
        (8,    6.4,   1.86, 18),
        (9,    5.4,   1.54, 16),
        (10,   4.6,   1.30, 15),
        (11,   3.9,   1.09, 14),
        (12,   3.3,   0.95, 13),
        (13,   2.7,   0.82, 12),
        (14,   2.3,   0.74, 11),
        (15,   2.0,   0.66, 10),
        (16,   1.65,  0.56, 10),    # Terminal bronchioles
        (17,   1.41,  0.45,  9),    # Respiratory bronchioles
        (18,   1.17,  0.39,  8),
        (19,   0.99,  0.33,  7),
        (20,   0.83,  0.27,  7),    # Alveolar ducts
        (21,   0.70,  0.23,  6),
        (22,   0.59,  0.20,  5),
        (23,   0.50,  0.18,  5),    # Alveolar sacs
    ]

    airways = []
    for gen, length_mm, diam_mm, angle_deg in weibel_data:
        airways.append(AirwayGeneration(
            generation=gen,
            length=length_mm * 1e-3,           # mm -> m
            diameter=diam_mm * 1e-3,            # mm -> m
            n_airways=2**gen,
            branching_angle=np.radians(angle_deg)
        ))
    return airways


# ============================================================================
# PARTICLE PHYSICS
# ============================================================================

def cunningham_slip(dp: float) -> float:
    """
    Cunningham slip correction factor for small particles.
    dp: particle diameter in meters
    """
    Kn = 2 * MEAN_FREE_PATH / dp  # Knudsen number
    return 1 + Kn * (2.34 + 1.05 * np.exp(-0.39 / Kn))


def diffusion_coefficient(dp: float) -> float:
    """
    Stokes-Einstein diffusion coefficient.
    dp: particle diameter in meters
    Returns: D in m²/s
    """
    Cc = cunningham_slip(dp)
    return (BOLTZMANN_K * TEMPERATURE * Cc) / (3 * np.pi * AIR_VISCOSITY * dp)


def settling_velocity(dp: float) -> float:
    """
    Terminal settling velocity under gravity (Stokes regime).
    dp: particle diameter in meters
    Returns: v_s in m/s
    """
    Cc = cunningham_slip(dp)
    return (PARTICLE_DENSITY * dp**2 * GRAVITY * Cc) / (18 * AIR_VISCOSITY)


def flow_velocity(airway: AirwayGeneration, tidal_volume: float, breathing_freq: float) -> float:
    """
    Mean air flow velocity in an airway generation.
    """
    # Volumetric flow rate (m³/s) — inspiratory flow
    Q = tidal_volume * breathing_freq / 60.0  # approximate
    # Cross-sectional area of all parallel airways
    A_total = airway.n_airways * np.pi * (airway.diameter / 2)**2
    if A_total == 0:
        return 0
    return Q / A_total


# ============================================================================
# DEPOSITION MECHANISMS
# ============================================================================

def deposition_brownian(dp: float, airway: AirwayGeneration, v_flow: float) -> float:
    """
    Deposition efficiency by Brownian diffusion.
    Uses Ingham (1975) formula for laminar flow in a cylindrical tube.
    """
    D = diffusion_coefficient(dp)
    R = airway.diameter / 2

    if v_flow <= 0 or R <= 0:
        return 0.0

    # Dimensionless deposition parameter (Delta)
    residence_time = airway.length / max(v_flow, 1e-10)
    delta = D * residence_time / (R**2)

    if delta > 0.1:
        # High diffusion regime
        eff = 1 - 0.819 * np.exp(-14.63 * delta) - 0.0976 * np.exp(-89.22 * delta) \
              - 0.0325 * np.exp(-228.0 * delta)
    else:
        # Low diffusion regime (Gormley & Kennedy)
        mu = (8 * delta) ** (1/3)
        eff = 1 - 0.819 * np.exp(-3.657 * mu) - 0.0976 * np.exp(-22.3 * mu)

    return np.clip(eff, 0, 1)


def deposition_sedimentation(dp: float, airway: AirwayGeneration, v_flow: float) -> float:
    """
    Deposition efficiency by gravitational sedimentation.
    Uses Pich (1972) formula for horizontal tubes.
    """
    v_s = settling_velocity(dp)
    R = airway.diameter / 2

    if v_flow <= 0 or R <= 0:
        return 0.0

    # Dimensionless settling parameter
    epsilon = (3 * v_s * airway.length) / (8 * R * v_flow)

    # Account for airway inclination (average gravity component)
    gravity_factor = np.cos(airway.branching_angle) if airway.branching_angle > 0 else 1.0
    epsilon *= gravity_factor

    if epsilon >= 1:
        return 1.0

    # Pich formula
    eff = (2 / np.pi) * (2 * epsilon * np.sqrt(1 - epsilon**(2/3))
                          - epsilon**(1/3) * np.sqrt(1 - epsilon**(2/3))
                          + np.arcsin(epsilon**(1/3)))

    return np.clip(eff, 0, 1)


def deposition_impaction(dp: float, airway: AirwayGeneration, v_flow: float) -> float:
    """
    Deposition efficiency by inertial impaction at bifurcations.
    Uses the Yeh & Schum (1980) empirical model.
    """
    Cc = cunningham_slip(dp)

    # Stokes number
    tau = (PARTICLE_DENSITY * dp**2 * Cc) / (18 * AIR_VISCOSITY)
    Stk = tau * v_flow / (airway.diameter / 2)

    if airway.branching_angle == 0:
        return 0.0  # No impaction at trachea (no bifurcation)

    # Empirical impaction formula (Zhang et al.)
    theta = airway.branching_angle
    eff = 1 - (2 / np.pi) * np.arccos(Stk * np.sin(theta))

    # Stk must be sufficient for impaction
    if Stk * np.sin(theta) > 1:
        eff = 1.0
    elif Stk * np.sin(theta) < 0:
        eff = 0.0

    return np.clip(eff, 0, 1)


# ============================================================================
# MONTE CARLO SIMULATION
# ============================================================================

def simulate_single_particle(
    dp_um: float,
    airways: List[AirwayGeneration],
    tidal_volume: float,
    breathing_freq: float,
    rng: np.random.Generator
) -> Dict[str, float]:
    """
    Simulate the journey of a single particle through the lung.

    Uses statistical weight method: instead of binary deposition,
    the particle's weight is reduced at each generation by the
    deposition probability.

    Returns: dict with regional deposition fractions
    """
    dp = dp_um * 1e-6  # convert µm to m
    weight = 1.0
    regional_deposition = {region: 0.0 for region in REGION_MAP}

    for airway in airways:
        if weight < 1e-10:
            break

        gen = airway.generation
        v_flow = flow_velocity(airway, tidal_volume, breathing_freq)

        # Add stochastic variation to airway geometry (±15% as in IDEAL-2)
        length_var = airway.length * (1 + 0.15 * rng.standard_normal())
        diam_var = airway.diameter * (1 + 0.10 * rng.standard_normal())

        # Create a varied airway for this particle
        varied_airway = AirwayGeneration(
            generation=gen,
            length=max(length_var, airway.length * 0.5),
            diameter=max(diam_var, airway.diameter * 0.5),
            n_airways=airway.n_airways,
            branching_angle=airway.branching_angle
        )

        # Calculate deposition by each mechanism
        p_diff = deposition_brownian(dp, varied_airway, v_flow)
        p_sed = deposition_sedimentation(dp, varied_airway, v_flow)
        p_imp = deposition_impaction(dp, varied_airway, v_flow)

        # Combined deposition probability (independent mechanisms)
        p_total = 1 - (1 - p_diff) * (1 - p_sed) * (1 - p_imp)
        p_total = np.clip(p_total, 0, 1)

        # Determine which region this generation belongs to
        for region, (gen_start, gen_end) in REGION_MAP.items():
            if gen_start <= gen < gen_end:
                deposited = weight * p_total
                regional_deposition[region] += deposited
                break

        # Reduce statistical weight
        weight *= (1 - p_total)

    # Remaining weight = exhaled (not deposited)
    return regional_deposition


def run_simulation(
    n_particles: int,
    age_group: str,
    age_groups: Dict[str, Dict[str, float | str]],
    pm_sizes: List[float],
    seed: int = DEFAULT_SEED,
    progress_callback: Optional[Callable[[Dict[str, object]], None]] = None,
    progress_offset: int = 0,
    progress_total: Optional[int] = None,
    log_callback: Optional[Callable[[str], None]] = None,
) -> Dict[float, Dict[str, Tuple[float, float]]]:
    """
    Run full Monte Carlo simulation for all PM sizes.

    Returns: {pm_size: {region: (mean_deposition%, std%)}}
    """
    if age_group not in age_groups:
        raise KeyError(f'Unknown age group {age_group!r}.')

    rng = np.random.default_rng(seed)
    airways = build_weibel_lung()
    tv = age_groups[age_group]['tidal_volume']
    bf = age_groups[age_group]['breathing_freq']

    results = {}
    update_interval = max(1, n_particles // 100)
    pm_count = len(pm_sizes)

    for pm_index, pm in enumerate(pm_sizes, start=1):
        region_deposits = {region: [] for region in REGION_MAP}

        for i in range(n_particles):
            dep = simulate_single_particle(pm, airways, tv, bf, rng)
            for region in REGION_MAP:
                region_deposits[region].append(dep[region] * 100)  # as percentage

            if progress_callback is not None and ((i + 1) % update_interval == 0 or i + 1 == n_particles):
                completed_age_units = (pm_index - 1) * n_particles + (i + 1)
                progress_callback({
                    'stage': 'simulation',
                    'age_group': age_group,
                    'pm_size': pm,
                    'pm_index': pm_index,
                    'pm_total': pm_count,
                    'particle': i + 1,
                    'particle_total': n_particles,
                    'completed_units': progress_offset + completed_age_units,
                    'total_units': progress_total,
                })

        results[pm] = {}
        for region in REGION_MAP:
            vals = np.array(region_deposits[region])
            results[pm][region] = (float(np.mean(vals)), float(np.std(vals)))

        total_dep = sum(results[pm][r][0] for r in REGION_MAP)
        emit_log(
            f"  PM {pm:>4.1f} um -> Total deposition: {total_dep:5.1f}% "
            f"(ET: {results[pm]['Extra-thoracic'][0]:5.1f}%, "
            f"Tubular: {results[pm]['Conducting (Tubular)'][0]:5.1f}%, "
            f"Alveolar: {results[pm]['Alveolar'][0]:5.1f}%)",
            log_callback,
        )

    return results


# ============================================================================
# MONTE CARLO FORMULA (from paper)
# ============================================================================

def monte_carlo_formula(n_particles: int, unit_density: float,
                        tidal_volume_cm3: float, breathing_freq: float) -> float:
    """
    Paper's formula:
    MC = (30 × N × ρ) / (TV × BF)
    For 8-hour average per day.
    """
    return (30 * n_particles * unit_density) / (tidal_volume_cm3 * breathing_freq)


# ============================================================================
# 3D VISUALIZATION (Fig 3.2 from paper)
# ============================================================================

def generate_3d_scatter(
    results: Dict,
    pm_sizes: List[float],
    age_label: Optional[str] = None,
    save_path: str = 'pm_3d_scatter.png',
    log_callback: Optional[Callable[[str], None]] = None,
):
    """
    Generate 3D scatter plot of PM distribution in alveolar system.
    Reproduces Fig 3.2 from the paper.
    """
    fig = plt.figure(figsize=(14, 10))
    ax = fig.add_subplot(111, projection='3d')

    colors = {
        0.1:  '#8B00FF',  # violet
        0.5:  '#0000FF',  # blue
        1.0:  '#00BFFF',  # cyan
        1.5:  '#00FF00',  # green
        2.0:  '#FFFF00',  # yellow
        2.5:  '#FF8C00',  # orange
        10.0: '#FF0000',  # red
    }

    labels = {
        0.1: 'PM 0.1 (Ultrafine)',
        0.5: 'PM 0.5',
        1.0: 'PM 1.0',
        1.5: 'PM 1.5',
        2.0: 'PM 2.0',
        2.5: 'PM 2.5 (Fine)',
        10.0: 'PM 10 (Coarse)',
    }

    rng = np.random.default_rng(123)
    palette = plt.cm.turbo(np.linspace(0.08, 0.92, len(pm_sizes)))
    fallback_colors = {
        pm: palette[index]
        for index, pm in enumerate(pm_sizes)
    }

    for pm in pm_sizes:
        # Number of scatter points proportional to alveolar deposition
        alv_dep = results[pm]['Alveolar'][0]
        n_points = max(int(alv_dep * 30), 20)

        # Generate lung-shaped distribution
        # Alveolar region is roughly spherical in the lower lung
        spread = alv_dep * 3 + 5
        theta = rng.uniform(0, 2 * np.pi, n_points)
        phi = rng.uniform(0, np.pi, n_points)
        r = rng.exponential(spread, n_points)

        x = r * np.sin(phi) * np.cos(theta)
        y = r * np.sin(phi) * np.sin(theta)
        z = r * np.cos(phi) - 10  # shift downward (lungs are below trachea)

        ax.scatter(
            x,
            y,
            z,
            c=[colors.get(pm, fallback_colors[pm])],
            s=8,
            alpha=0.6,
            label=labels.get(pm, f'PM {pm}'),
        )

    ax.set_xlabel('X (mm)', fontsize=10)
    ax.set_ylabel('Y (mm)', fontsize=10)
    ax.set_zlabel('Z (depth, mm)', fontsize=10)
    title = '3D Particulate Spread in Alveolar System\n(Monte Carlo Simulation - Multiple Runs)'
    if age_label:
        title += f'\nAge Group: {age_label}'
    ax.set_title(title, fontsize=13, fontweight='bold')
    ax.legend(loc='upper left', fontsize=9, markerscale=3)
    ax.view_init(elev=25, azim=45)

    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches='tight')
    plt.close()
    emit_log(f"  -> Saved 3D scatter plot: {save_path}", log_callback)


# ============================================================================
# BAR CHART VISUALIZATION (Fig 4.1 from paper)
# ============================================================================

def generate_deposition_bar_charts(
    all_age_results: Dict,
    pm_sizes: List[float],
    age_groups: Dict[str, AgeGroupConfig],
    n_particles: int,
    save_path: str = 'pm_deposition_bars.png',
    log_callback: Optional[Callable[[str], None]] = None,
):
    """
    Generate grouped bar charts showing PM deposition % by region and age group.
    Reproduces Fig 4.1 from the paper.
    """
    fig, axes = plt.subplots(1, 3, figsize=(18, 7), sharey=False)
    regions = list(REGION_MAP.keys())
    region_titles = ['Deposition in Extra-thoracic Region',
                     'Deposition in Conducting Airways',
                     'Deposition in Alveolar Region']

    bar_colors = ['#2ecc71', '#3498db', '#e74c3c', '#f39c12']
    bar_width = 0.18

    for idx, (region, title) in enumerate(zip(regions, region_titles)):
        ax = axes[idx]
        x = np.arange(len(pm_sizes))

        for i, (age_key, age_data) in enumerate(all_age_results.items()):
            means = [age_data[pm][region][0] for pm in pm_sizes]
            label = age_groups[age_key].label
            if 'yr' not in label.lower():
                label = f'{label} yrs'
            ax.bar(x + i * bar_width, means, bar_width,
                   label=label, color=bar_colors[i % len(bar_colors)], alpha=0.85, edgecolor='white')

        ax.set_xlabel('Particle Size (µm)', fontsize=11)
        ax.set_ylabel('Deposition (%)', fontsize=11)
        ax.set_title(title, fontsize=12, fontweight='bold')
        ax.set_xticks(x + bar_width * 1.5)
        ax.set_xticklabels([f'{pm}' for pm in pm_sizes], fontsize=9)
        ax.legend(fontsize=8, loc='upper right')
        ax.grid(axis='y', alpha=0.3)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

    plt.suptitle(f'Percent Occupancy of PM in Respiratory System by Age Group\n(Monte Carlo Simulation, N={n_particles:,} particles)',
                 fontsize=14, fontweight='bold', y=1.02)
    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches='tight')
    plt.close()
    emit_log(f"  -> Saved deposition bar charts: {save_path}", log_callback)


# ============================================================================
# DEPOSITION HEATMAP
# ============================================================================

def generate_heatmap(
    results: Dict,
    age_label: str,
    pm_sizes: List[float],
    save_path: str = 'pm_heatmap.png',
    log_callback: Optional[Callable[[str], None]] = None,
):
    """
    Generate a heatmap of deposition % by PM size and lung region.
    """
    regions = list(REGION_MAP.keys())
    region_short = ['Extra-thoracic', 'Tubular', 'Alveolar']
    data = np.zeros((len(pm_sizes), len(regions)))

    for i, pm in enumerate(pm_sizes):
        for j, region in enumerate(regions):
            data[i, j] = results[pm][region][0]

    fig, ax = plt.subplots(figsize=(10, 7))
    im = ax.imshow(data, cmap='YlOrRd', aspect='auto')

    ax.set_xticks(range(len(regions)))
    ax.set_xticklabels(region_short, fontsize=11)
    ax.set_yticks(range(len(pm_sizes)))
    ax.set_yticklabels([f'PM {pm}' for pm in pm_sizes], fontsize=11)

    # Annotate cells
    for i in range(len(pm_sizes)):
        for j in range(len(regions)):
            val = data[i, j]
            color = 'white' if val > data.max() * 0.6 else 'black'
            ax.text(j, i, f'{val:.1f}%', ha='center', va='center',
                    fontsize=11, fontweight='bold', color=color)

    ax.set_title(f'PM Deposition (%) in Respiratory Regions — {age_label}',
                 fontsize=13, fontweight='bold', pad=15)
    plt.colorbar(im, ax=ax, label='Deposition %', shrink=0.8)

    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches='tight')
    plt.close()
    emit_log(f"  -> Saved heatmap: {save_path}", log_callback)


# ============================================================================
# LUNG INFOGRAPHIC — ANNOTATED ANATOMICAL PANEL
# ============================================================================

def generate_lung_infographic(
    results: Dict,
    age_label: str,
    pm_sizes: List[float],
    pm_size: Optional[float] = None,
    save_path: str = 'pm_lung_infographic.png',
    log_callback: Optional[Callable[[str], None]] = None,
):
    """
    Generate an annotated lung infographic showing PM deposition by region.
    Left panel: stylized anatomical lung diagram with color-coded regions.
    Right panel: horizontal bar chart of deposition fractions.
    """
    if not pm_sizes:
        raise ValueError('At least one PM size is required to generate the infographic.')
    if pm_size is None:
        pm_size = 2.5 if 2.5 in pm_sizes else pm_sizes[0]

    et_dep = results[pm_size]['Extra-thoracic'][0]
    tb_dep = results[pm_size]['Conducting (Tubular)'][0]
    al_dep = results[pm_size]['Alveolar'][0]
    total_dep = et_dep + tb_dep + al_dep

    # Color palette — darker = higher deposition
    cmap = plt.cm.YlOrRd
    max_dep = max(et_dep, tb_dep, al_dep)
    et_color = cmap(et_dep / max_dep * 0.85 + 0.1)
    tb_color = cmap(tb_dep / max_dep * 0.85 + 0.1)
    al_color = cmap(al_dep / max_dep * 0.85 + 0.1)

    fig = plt.figure(figsize=(18, 11), facecolor='#f8f9fa')
    fig.suptitle(
        f'PM {pm_size} µm Deposition in Human Respiratory System — {age_label}',
        fontsize=17, fontweight='bold', y=0.97, color='#2c3e50'
    )

    # ----------------------------------------------------------------
    # LEFT PANEL — Anatomical lung schematic
    # ----------------------------------------------------------------
    ax_lung = fig.add_axes([0.02, 0.05, 0.52, 0.85])
    ax_lung.set_xlim(-0.5, 10.5)
    ax_lung.set_ylim(-0.5, 15.5)
    ax_lung.set_aspect('equal')
    ax_lung.axis('off')

    # Anatomical colour palette
    lung_fill = '#fadbd8'
    lung_fill2 = '#f5b7b1'
    lung_edge = '#943126'
    airway_wall = '#a93226'
    cartilage_c = '#d5dbdb'
    mucosa_c = '#f1948a'

    rng = np.random.default_rng(42)
    alveoli_positions = []

    # ---- Helper: recursive bronchial tree ----
    def _draw_tree(x, y, angle, length, width, depth, max_depth):
        if depth > max_depth or width < 0.15:
            alveoli_positions.append((x, y))
            return
        rad = np.radians(angle)
        ex = x + length * np.sin(rad)
        ey = y - length * np.cos(rad)
        ax_lung.plot([x, ex], [y, ey], color=tb_color,
                     linewidth=width, solid_capstyle='round', alpha=0.9, zorder=4)
        if width > 1.0:
            ax_lung.plot([x, ex], [y, ey], color=mucosa_c,
                         linewidth=width * 0.35, solid_capstyle='round',
                         alpha=0.35, zorder=5)
        spread = max(10, 28 - depth * 3 + rng.uniform(-4, 4))
        nl = length * (0.72 + rng.uniform(-0.04, 0.04))
        nw = width * 0.62
        _draw_tree(ex, ey, angle - spread, nl, nw, depth + 1, max_depth)
        _draw_tree(ex, ey, angle + spread, nl, nw, depth + 1, max_depth)

    # ======== LUNG SILHOUETTES (background, zorder=1) ========
    # Diaphragm sweeping curve under the lungs for realism
    diaph_v = [(-0.5, 0.0), (1.0, 1.2), (4.5, 1.1), (5.0, 0.8), (5.5, 1.1), (9.0, 1.2), (10.5, 0.0)]
    diaph_c = [Path.MOVETO, Path.CURVE4, Path.CURVE4, Path.CURVE4, Path.CURVE4, Path.CURVE4, Path.CURVE4]
    ax_lung.add_patch(PathPatch(Path(diaph_v, diaph_c), facecolor='none', edgecolor='#d35400', linewidth=3, alpha=0.4, zorder=0))
    ax_lung.text(10.2, 0.4, 'Diaphragm', ha='right', va='center',
                 fontsize=7.5, fontweight='bold', color='#d35400', alpha=0.7, zorder=0)

    # Left lung — 2 lobes with cardiac notch (7 cubic Bézier segments)
    ll = [
        (4.0, 9.0),
        (3.2, 9.55), (1.9, 9.45), (1.0, 8.7),
        (0.4, 7.8), (0.15, 6.5), (0.15, 5.3),
        (0.15, 3.8), (0.4, 2.3), (1.1, 1.4),
        (1.9, 0.9), (3.1, 0.9), (3.8, 1.4),
        (4.15, 2.1), (4.3, 3.4), (4.15, 4.5),
        (3.85, 5.1), (3.65, 5.55), (3.7, 6.1),
        (3.85, 6.7), (4.1, 7.8), (4.0, 9.0),
        (4.0, 9.0),
    ]
    lc = [Path.MOVETO] + [Path.CURVE4] * 21 + [Path.CLOSEPOLY]
    ax_lung.add_patch(PathPatch(Path(ll, lc), facecolor=lung_fill,
                                edgecolor=lung_edge, linewidth=2.2, alpha=0.85, zorder=1))
    ax_lung.add_patch(Ellipse((2.0, 5.0), 2.8, 6.0, facecolor=lung_fill2,
                               edgecolor='none', alpha=0.3, zorder=1))

    # Right lung — 3 lobes, wider (6 cubic Bézier segments)
    rl = [
        (6.0, 9.0),
        (6.8, 9.55), (8.1, 9.45), (9.0, 8.7),
        (9.6, 7.8), (9.85, 6.5), (9.85, 5.3),
        (9.85, 3.8), (9.6, 2.3), (8.9, 1.4),
        (8.1, 0.9), (6.9, 0.9), (6.2, 1.4),
        (5.85, 2.1), (5.75, 3.5), (5.8, 5.0),
        (5.8, 6.2), (5.85, 7.8), (6.0, 9.0),
        (6.0, 9.0),
    ]
    rc = [Path.MOVETO] + [Path.CURVE4] * 18 + [Path.CLOSEPOLY]
    ax_lung.add_patch(PathPatch(Path(rl, rc), facecolor=lung_fill,
                                edgecolor=lung_edge, linewidth=2.2, alpha=0.85, zorder=1))
    ax_lung.add_patch(Ellipse((8.0, 5.0), 2.8, 6.0, facecolor=lung_fill2,
                               edgecolor='none', alpha=0.3, zorder=1))

    # ---- Lobe fissures (dashed lines) ----
    ax_lung.plot([3.7, 0.7], [8.0, 2.0], color=lung_edge, linewidth=1.2,
                 linestyle='--', alpha=0.45, zorder=2)
    ax_lung.plot([6.3, 9.3], [8.0, 2.0], color=lung_edge, linewidth=1.2,
                 linestyle='--', alpha=0.45, zorder=2)
    ax_lung.plot([6.0, 9.5], [6.3, 6.3], color=lung_edge, linewidth=1.2,
                 linestyle='--', alpha=0.45, zorder=2)

    # ======== TRACHEA ========
    trach_cx, trach_cy, trach_w, trach_h = 5.0, 10.4, 1.0, 2.0
    # Trachea tube
    ax_lung.add_patch(FancyBboxPatch(
        (trach_cx - trach_w / 2, trach_cy - trach_h / 2), trach_w, trach_h,
        boxstyle="round,pad=0.1", facecolor=tb_color, edgecolor=airway_wall,
        linewidth=1.8, alpha=0.92, zorder=6))
    
    # Improved 3D-looking cartilage rings
    for ry in np.linspace(trach_cy - trach_h / 2 + 0.15,
                          trach_cy + trach_h / 2 - 0.15, 8):
        ax_lung.add_patch(Arc((trach_cx, ry), trach_w * 0.9, 0.25, angle=0,
                              theta1=180, theta2=360, color=cartilage_c,
                              linewidth=2.5, alpha=0.9, zorder=7))
        ax_lung.add_patch(Arc((trach_cx, ry - 0.05), trach_w * 0.9, 0.25, angle=0,
                              theta1=180, theta2=360, color='white',
                              linewidth=1, alpha=0.4, zorder=7)) # highlight
    
    ax_lung.text(6.1, 10.4, 'Trachea', ha='left', va='center',
                 fontsize=8.5, fontweight='bold', color='#2c3e50', zorder=8)

    # ======== PHARYNX / LARYNX ========
    # More realistic sweeping shape for larynx (thyroid cartilage / funnel)
    larynx_verts = [
        (4.4, 11.5),   # bottom left (connects to trachea)
        (4.1, 12.2),   # flare out left
        (4.2, 12.8),   # straight left pharynx
        (4.3, 13.4),   # curve towards nasal
        (5.7, 13.4),   # across to right nasal
        (5.8, 12.8),   # straight right pharynx
        (5.9, 12.2),   # flare out right
        (5.6, 11.5),   # bottom right
        (4.4, 11.5)    # close
    ]
    larynx_codes = [Path.MOVETO, Path.CURVE4, Path.CURVE4, Path.CURVE4, 
                    Path.LINETO, Path.CURVE4, Path.CURVE4, Path.CURVE4, Path.CLOSEPOLY]
    # Smooth the path with Spline-like bezier for realism
    ax_lung.add_patch(PathPatch(Path(larynx_verts, larynx_codes), facecolor=et_color,
                                edgecolor=airway_wall, linewidth=1.8, alpha=0.9, zorder=5))
    
    # Anatomical Larynx lines (Thyroid cartilage & Cricoid)
    ax_lung.plot([4.6, 5.0, 5.4], [11.8, 11.6, 11.8], color=airway_wall, linewidth=1.5, alpha=0.6, zorder=6) # Cricoid
    ax_lung.plot([4.3, 5.0, 5.7], [12.2, 11.9, 12.2], color=airway_wall, linewidth=1.5, alpha=0.6, zorder=6) # Thyroid base
    ax_lung.plot([5.0, 5.0], [11.9, 12.5], color=airway_wall, linewidth=1.5, alpha=0.6, zorder=6) # Prominence line
    
    ax_lung.text(6.1, 12.0, 'Larynx', ha='left', va='center',
                 fontsize=8.0, fontweight='bold', color='#2c3e50', zorder=8)
    ax_lung.text(6.1, 12.8, 'Pharynx', ha='left', va='center',
                 fontsize=8.0, fontweight='bold', color='#2c3e50', zorder=8)

    # ======== CONTINUOUS BRONCHIAL TREE (Upper parts of lungs connected to Trachea) ========
    # We replace the hardcoded square carina connection with a realistic split
    carina_y = trach_cy - trach_h / 2
    # Left main bronchus curve
    l_bronchus = [(5.0, carina_y), (4.7, carina_y - 0.4), (3.8, carina_y - 0.8), (3.3, carina_y - 1.2)]
    ax_lung.add_patch(PathPatch(Path(l_bronchus, [Path.MOVETO, Path.CURVE4, Path.CURVE4, Path.CURVE4]), 
                                facecolor='none', edgecolor=tb_color, linewidth=18, 
                                capstyle='round', alpha=0.92, zorder=5))
    # Right main bronchus curve (steeper, wider anatomically)
    r_bronchus = [(5.0, carina_y), (5.4, carina_y - 0.3), (6.2, carina_y - 0.9), (6.7, carina_y - 1.4)]
    ax_lung.add_patch(PathPatch(Path(r_bronchus, [Path.MOVETO, Path.CURVE4, Path.CURVE4, Path.CURVE4]), 
                                facecolor='none', edgecolor=tb_color, linewidth=22, 
                                capstyle='round', alpha=0.92, zorder=5))
    
    # Add cartilage rings to main bronchi
    for t in np.linspace(0.2, 0.8, 4):
        lx, ly = 5.0*(1-t)**3 + 3*4.7*t*(1-t)**2 + 3*3.8*t**2*(1-t) + 3.3*t**3, carina_y*(1-t)**3 + 3*(carina_y-0.4)*t*(1-t)**2 + 3*(carina_y-0.8)*t**2*(1-t) + (carina_y-1.2)*t**3
        rx, ry = 5.0*(1-t)**3 + 3*5.4*t*(1-t)**2 + 3*6.2*t**2*(1-t) + 6.7*t**3, carina_y*(1-t)**3 + 3*(carina_y-0.3)*t*(1-t)**2 + 3*(carina_y-0.9)*t**2*(1-t) + (carina_y-1.4)*t**3
        # Left rings
        ax_lung.add_patch(Arc((lx, ly), 0.7, 0.2, angle=30, theta1=180, theta2=360, color=cartilage_c, linewidth=2, alpha=0.8, zorder=6))
        # Right rings
        ax_lung.add_patch(Arc((rx, ry), 0.85, 0.25, angle=-35, theta1=180, theta2=360, color=cartilage_c, linewidth=2, alpha=0.8, zorder=6))

    # ======== NASAL CAVITY (anatomical coronal section) ========
    # Smooth, realistic nasal cavity silhouette replacing the "external nose"
    nose_outer = [
        (5.0, 14.8),                                         # Septum top
        (4.5, 14.8), (4.2, 14.5), (4.1, 14.1),               # left roof and lateral wall
        (4.1, 13.8), (4.3, 13.5), (4.4, 13.4),               # left curve in
        (4.3, 13.3), (4.5, 13.2), (4.6, 13.3),               # left lowest turbinate curve (bottom)
        (4.8, 13.4), (4.9, 13.4), (5.0, 13.4),               # meet septum base
        (5.1, 13.4), (5.2, 13.4), (5.4, 13.3),               # right septum base
        (5.5, 13.2), (5.7, 13.3), (5.6, 13.4),               # right lowest turbinate curve (bottom)
        (5.7, 13.5), (5.9, 13.8), (5.9, 14.1),               # right lateral wall
        (5.8, 14.5), (5.5, 14.8), (5.0, 14.8),               # right roof to septum
        (5.0, 14.8),
    ]
    nose_codes = [Path.MOVETO] + [Path.CURVE4] * 24 + [Path.CLOSEPOLY]
    ax_lung.add_patch(PathPatch(Path(nose_outer, nose_codes), facecolor=et_color,
                                edgecolor=airway_wall, linewidth=1.8, alpha=0.9, zorder=6))
    # Nasal septum (vertical midline)
    ax_lung.plot([5.0, 5.0], [14.35, 13.1], color=airway_wall,
                 linewidth=1.2, alpha=0.5, zorder=7)
    # Turbinate ridges (three curved shelves per side)
    for ty, tw in [(14.05, 0.28), (13.7, 0.35), (13.35, 0.30)]:
        # left turbinate
        ax_lung.add_patch(Arc((4.65, ty), tw, 0.12, angle=0,
                              theta1=180, theta2=360, color='#d98880',
                              linewidth=1.6, alpha=0.6, zorder=7))
        # right turbinate
        ax_lung.add_patch(Arc((5.35, ty), tw, 0.12, angle=0,
                              theta1=180, theta2=360, color='#d98880',
                              linewidth=1.6, alpha=0.6, zorder=7))
    # Nostril openings
    ax_lung.add_patch(Ellipse((4.65, 12.95), 0.18, 0.12, facecolor='#2c3e50',
                               edgecolor='none', alpha=0.55, zorder=7))
    ax_lung.add_patch(Ellipse((5.35, 12.95), 0.18, 0.12, facecolor='#2c3e50',
                               edgecolor='none', alpha=0.55, zorder=7))
    ax_lung.text(5, 14.15, 'Nasal', ha='center', va='center',
                 fontsize=6.5, fontweight='bold', color='#2c3e50', zorder=8)
    ax_lung.text(5, 13.95, 'Cavity', ha='center', va='center',
                 fontsize=6.5, fontweight='bold', color='#2c3e50', zorder=8)

    # ======== BRONCHIAL TREES ========
    # Start trees from the ends of the new main bronchi we drew
    _draw_tree(3.3, carina_y - 1.2, -15, 1.5, 3.2, 0, 5)
    _draw_tree(6.7, carina_y - 1.4,  15, 1.5, 3.2, 0, 5)

    # ======== ALVEOLAR CLUSTERS ========
    for (acx, acy) in alveoli_positions:
        if acx < 0.1 or acx > 9.9 or acy < 0.8 or acy > 8.8:
            continue
        for _ in range(rng.integers(4, 8)):
            dx, dy = rng.uniform(-0.22, 0.22), rng.uniform(-0.22, 0.22)
            r = rng.uniform(0.07, 0.16)
            ax_lung.add_patch(Ellipse((acx + dx, acy + dy), r * 2, r * 2,
                                       facecolor=al_color, edgecolor='#c0392b',
                                       linewidth=0.4, alpha=0.65, zorder=3))

    # ======== PM DEPOSITION SPOTS ========
    # Scatter coloured dots proportional to deposition in each region
    dep_rng = np.random.default_rng(99)
    spot_cmap = plt.cm.hot_r  # warm colours for deposited particles

    # --- ET region spots (nasal/pharynx area) ---
    n_et = max(3, int(et_dep * 0.6))
    for _ in range(n_et):
        sx = dep_rng.uniform(4.25, 5.75)
        sy = dep_rng.uniform(12.9, 14.3)
        sr = dep_rng.uniform(0.04, 0.09)
        ax_lung.add_patch(Ellipse((sx, sy), sr * 2, sr * 2,
                                   facecolor='#e74c3c', edgecolor='none',
                                   alpha=0.75, zorder=9))

    # --- Tubular region spots (trachea & bronchi) ---
    n_tb = max(5, int(tb_dep * 0.5))
    for _ in range(n_tb):
        sx = dep_rng.uniform(1.5, 8.5)
        sy = dep_rng.uniform(5.0, 9.5)
        # keep within lung outlines roughly
        if sx < 5:
            if sx < 0.5 or sx > 4.2:
                continue
        else:
            if sx > 9.5 or sx < 5.8:
                continue
        sr = dep_rng.uniform(0.05, 0.12)
        ax_lung.add_patch(Ellipse((sx, sy), sr * 2, sr * 2,
                                   facecolor='#c0392b', edgecolor='none',
                                   alpha=0.7, zorder=9))

    # --- Alveolar region spots (lower lung periphery) ---
    n_al = max(3, int(al_dep * 0.8))
    for _ in range(n_al):
        sx = dep_rng.uniform(0.6, 9.4)
        sy = dep_rng.uniform(1.2, 5.0)
        # keep within lung silhouettes
        if sx < 5:
            if sx < 0.4 or sx > 4.1:
                continue
        else:
            if sx > 9.6 or sx < 5.9:
                continue
        sr = dep_rng.uniform(0.06, 0.14)
        ax_lung.add_patch(Ellipse((sx, sy), sr * 2, sr * 2,
                                   facecolor='#922b21', edgecolor='none',
                                   alpha=0.65, zorder=9))

    # Deposition legend
    ax_lung.scatter([], [], c='#e74c3c', s=30, label=f'ET deposit ({et_dep:.1f}%)')
    ax_lung.scatter([], [], c='#c0392b', s=30, label=f'Tubular deposit ({tb_dep:.1f}%)')
    ax_lung.scatter([], [], c='#922b21', s=30, label=f'Alveolar deposit ({al_dep:.1f}%)')
    ax_lung.legend(loc='lower right', fontsize=7.5, framealpha=0.9,
                   edgecolor='#bdc3c7', fancybox=True, handletextpad=0.4)

    # ======== REGION LABELS ========
    ax_lung.annotate('EXTRA-THORACIC\nREGION', xy=(4.2, 13.2), xytext=(0.5, 13.8),
                     fontsize=9, fontweight='bold', color='#d35400',
                     arrowprops=dict(arrowstyle='->', color='#d35400', lw=1.5),
                     bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                               edgecolor='#d35400', alpha=0.9), zorder=10)
    ax_lung.annotate('CONDUCTING\nAIRWAYS', xy=(5.8, 8.5), xytext=(8.2, 11.0),
                     fontsize=9, fontweight='bold', color='#a93226',
                     arrowprops=dict(arrowstyle='->', color='#a93226', lw=1.5),
                     bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                               edgecolor='#a93226', alpha=0.9), zorder=10)
    ax_lung.annotate('ALVEOLAR\nREGION', xy=(2.2, 4.0), xytext=(0, 0.5),
                     fontsize=9, fontweight='bold', color='#c0392b',
                     arrowprops=dict(arrowstyle='->', color='#c0392b', lw=1.5),
                     bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                               edgecolor='#c0392b', alpha=0.9), zorder=10)

    # Airflow arrow
    ax_lung.annotate('', xy=(5, 14.6), xytext=(5, 15.2),
                     arrowprops=dict(arrowstyle='->', color='#3498db', lw=2.5),
                     zorder=10)
    ax_lung.text(5, 15.3, 'Inhaled Air + PM', ha='center', va='bottom',
                 fontsize=9, color='#3498db', fontweight='bold', zorder=10)

    # ----------------------------------------------------------------
    # RIGHT PANEL — Data bars + comparison across all PM sizes
    # ----------------------------------------------------------------
    ax_bar = fig.add_axes([0.58, 0.42, 0.38, 0.45])

    regions_display = ['Extra-thoracic', 'Conducting\n(Tubular)', 'Alveolar']
    dep_values = [et_dep, tb_dep, al_dep]
    bar_colors = [et_color, tb_color, al_color]

    bars = ax_bar.barh(regions_display, dep_values, color=bar_colors,
                       edgecolor='#7f8c8d', linewidth=1, height=0.55, alpha=0.9)

    for bar, val in zip(bars, dep_values):
        ax_bar.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height() / 2,
                    f'{val:.1f}%', va='center', fontsize=12, fontweight='bold',
                    color='#2c3e50')

    ax_bar.set_xlim(0, max(dep_values) * 1.2)
    ax_bar.set_xlabel('Deposition (%)', fontsize=11, color='#2c3e50')
    ax_bar.set_title(f'Regional Deposition — PM {pm_size} µm',
                     fontsize=13, fontweight='bold', color='#2c3e50', pad=10)
    ax_bar.spines['top'].set_visible(False)
    ax_bar.spines['right'].set_visible(False)
    ax_bar.tick_params(labelsize=10)
    ax_bar.grid(axis='x', alpha=0.3)

    # Total deposition badge
    ax_bar.text(max(dep_values) * 0.55, -0.7,
                f'Total Deposition: {total_dep:.1f}%',
                fontsize=12, fontweight='bold', color='white',
                bbox=dict(boxstyle='round,pad=0.4', facecolor='#e74c3c', alpha=0.9),
                ha='center')

    # ----------------------------------------------------------------
    # BOTTOM RIGHT — Small multi-PM comparison table
    # ----------------------------------------------------------------
    ax_table = fig.add_axes([0.58, 0.06, 0.38, 0.30])
    ax_table.axis('off')
    ax_table.set_title('Deposition Across All PM Sizes',
                       fontsize=12, fontweight='bold', color='#2c3e50', pad=8)

    col_labels = ['PM (µm)', 'ET (%)', 'Tubular (%)', 'Alveolar (%)', 'Total (%)']
    table_data = []
    for pm in pm_sizes:
        et = results[pm]['Extra-thoracic'][0]
        tb = results[pm]['Conducting (Tubular)'][0]
        al = results[pm]['Alveolar'][0]
        table_data.append([f'{pm}', f'{et:.1f}', f'{tb:.1f}', f'{al:.1f}',
                           f'{et + tb + al:.1f}'])

    table = ax_table.table(cellText=table_data, colLabels=col_labels,
                           loc='center', cellLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1, 1.4)

    # Style header
    for j in range(len(col_labels)):
        table[0, j].set_facecolor('#2c3e50')
        table[0, j].set_text_props(color='white', fontweight='bold')

    # Highlight the selected PM row
    pm_idx = pm_sizes.index(pm_size) + 1  # +1 for header
    for j in range(len(col_labels)):
        table[pm_idx, j].set_facecolor('#ffeaa7')
        table[pm_idx, j].set_edgecolor('#f39c12')

    # Alternate row shading
    for i in range(1, len(pm_sizes) + 1):
        if i != pm_idx and i % 2 == 0:
            for j in range(len(col_labels)):
                table[i, j].set_facecolor('#f0f0f0')

    plt.savefig(save_path, dpi=180, bbox_inches='tight', facecolor=fig.get_facecolor())
    plt.close()
    emit_log(f"  -> Saved lung infographic: {save_path}", log_callback)


# ============================================================================
# RESULTS TABLE
# ============================================================================

def render_results_table(results: Dict, age_label: str, pm_sizes: List[float]) -> str:
    """Render a formatted results table matching Table 4.1 from the paper."""
    lines = [
        '',
        '=' * 70,
        f'  PM Deposition Results - {age_label}',
        '=' * 70,
        f"  {'PM (um)':<10} {'Extra-thoracic':>16} {'Tubular':>16} {'Alveolar':>16}",
        f"  {'-' * 58}",
    ]

    for pm in pm_sizes:
        et = results[pm]['Extra-thoracic']
        tb = results[pm]['Conducting (Tubular)']
        al = results[pm]['Alveolar']
        lines.append(
            f"  PM {pm:<6.1f} {et[0]:>11.1f}% +/-{et[1]:<4.1f} "
            f"{tb[0]:>11.1f}% +/-{tb[1]:<4.1f} "
            f"{al[0]:>11.1f}% +/-{al[1]:<4.1f}"
        )

    lines.extend(['=' * 70, ''])
    return '\n'.join(lines)


def print_results_table(
    results: Dict,
    age_label: str,
    pm_sizes: List[float],
    log_callback: Optional[Callable[[str], None]] = None,
):
    """Print or forward a formatted results table."""
    emit_log(render_results_table(results, age_label, pm_sizes), log_callback)


# ============================================================================
# WORKBOOK IO AND APPLICATION ORCHESTRATION
# ============================================================================

def validate_integer(value, field_name: str, *, allow_zero: bool = False) -> int:
    """Parse integer workbook cells while rejecting fractional values."""
    numeric_value = validate_numeric(value, field_name, allow_zero=allow_zero)
    if not float(numeric_value).is_integer():
        raise ValueError(f'{field_name} must be a whole number, got {numeric_value}.')
    return int(numeric_value)


def format_age_display(age_group: AgeGroupConfig) -> str:
    """Use a readable label in the UI and in generated charts."""
    return age_group.label if 'yr' in age_group.label.lower() else f'{age_group.label} yrs'


def slugify_for_filename(value: str) -> str:
    """Create a filesystem-safe stem from an age-group identifier."""
    slug = ''.join(character.lower() if character.isalnum() else '_' for character in value.strip())
    slug = '_'.join(part for part in slug.split('_') if part)
    return slug or 'age_group'


def report_progress(
    progress_callback: Optional[Callable[[Dict[str, object]], None]],
    *,
    stage: str,
    message: str,
    completed_units: int,
    total_units: Optional[int],
    **extra,
):
    """Send a normalized progress payload to the UI or CLI caller."""
    if progress_callback is None:
        return

    payload: Dict[str, object] = {
        'stage': stage,
        'message': message,
        'completed_units': completed_units,
        'total_units': total_units,
    }
    if total_units:
        payload['progress'] = max(0.0, min(1.0, completed_units / total_units))
    payload.update(extra)
    progress_callback(payload)


def create_template_workbook(
    file_path: str | FilePath,
    config: Optional[SimulationConfig] = None,
):
    """Create a starter workbook that mirrors the script's default data."""
    config = clone_config(config or get_default_config())
    target_path = FilePath(file_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    settings_sheet = workbook.active
    settings_sheet.title = WORKBOOK_SHEETS['settings']
    settings_sheet.append(['key', 'value', 'description'])
    settings_sheet.append([
        'n_particles',
        config.n_particles,
        'Particles simulated per PM size for each age group.',
    ])
    settings_sheet.append([
        'seed',
        config.seed,
        'Random seed used to reproduce stochastic airway perturbations.',
    ])

    pm_sheet = workbook.create_sheet(WORKBOOK_SHEETS['pm_sizes'])
    pm_sheet.append(['pm_size_um'])
    for pm_size in config.pm_sizes:
        pm_sheet.append([pm_size])

    age_sheet = workbook.create_sheet(WORKBOOK_SHEETS['age_groups'])
    age_sheet.append(['group_key', 'label', 'tidal_volume_cm3', 'breathing_freq_per_min'])
    for age_key, age_group in config.age_groups.items():
        age_sheet.append([
            age_key,
            age_group.label,
            age_group.tidal_volume_cm3,
            age_group.breathing_freq,
        ])

    workbook.save(target_path)


def read_sheet_headers(worksheet) -> Dict[str, int]:
    """Return a header-to-index map for a workbook sheet."""
    headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if headers is None:
        raise ValueError(f'Sheet {worksheet.title!r} is empty.')

    return {
        str(header).strip(): index
        for index, header in enumerate(headers)
        if header is not None and str(header).strip()
    }


def load_config_from_workbook(file_path: str | FilePath) -> SimulationConfig:
    """Load and validate a simulation workbook."""
    source_path = FilePath(file_path)
    if not source_path.exists():
        raise FileNotFoundError(f'Workbook not found: {source_path}')

    workbook = load_workbook(source_path, data_only=True)
    try:
        missing_sheets = [
            sheet_name
            for sheet_name in WORKBOOK_SHEETS.values()
            if sheet_name not in workbook.sheetnames
        ]
        if missing_sheets:
            raise ValueError(f'Missing workbook sheets: {", ".join(missing_sheets)}.')

        settings_sheet = workbook[WORKBOOK_SHEETS['settings']]
        settings = {}
        for row in settings_sheet.iter_rows(min_row=2, values_only=True):
            key = row[0] if len(row) > 0 else None
            value = row[1] if len(row) > 1 else None
            if key is None or str(key).strip() == '':
                continue
            settings[str(key).strip()] = value

        n_particles = validate_integer(settings.get('n_particles'), 'n_particles')
        seed_value = settings.get('seed', DEFAULT_SEED)
        seed = validate_integer(seed_value if seed_value is not None else DEFAULT_SEED, 'seed', allow_zero=True)

        pm_sheet = workbook[WORKBOOK_SHEETS['pm_sizes']]
        pm_headers = read_sheet_headers(pm_sheet)
        if 'pm_size_um' not in pm_headers:
            raise ValueError(f"Sheet {pm_sheet.title!r} must contain a 'pm_size_um' column.")
        pm_sizes = []
        pm_index = pm_headers['pm_size_um']
        for row_number, row in enumerate(pm_sheet.iter_rows(min_row=2, values_only=True), start=2):
            value = row[pm_index] if pm_index < len(row) else None
            if value is None or str(value).strip() == '':
                continue
            pm_sizes.append(validate_numeric(value, f'pm_size_um (row {row_number})'))
        if not pm_sizes:
            raise ValueError('The workbook must include at least one PM size.')

        age_sheet = workbook[WORKBOOK_SHEETS['age_groups']]
        age_headers = read_sheet_headers(age_sheet)
        required_age_headers = ['group_key', 'label', 'tidal_volume_cm3', 'breathing_freq_per_min']
        missing_age_headers = [header for header in required_age_headers if header not in age_headers]
        if missing_age_headers:
            raise ValueError(
                f"Sheet {age_sheet.title!r} is missing required columns: {', '.join(missing_age_headers)}."
            )

        age_groups: Dict[str, AgeGroupConfig] = {}
        for row_number, row in enumerate(age_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            group_key = row[age_headers['group_key']]
            if group_key is None or str(group_key).strip() == '':
                raise ValueError(f'group_key is required in age_groups row {row_number}.')
            group_key = str(group_key).strip()
            if group_key in age_groups:
                raise ValueError(f'Duplicate age-group key {group_key!r} in row {row_number}.')

            label = normalize_age_label(row[age_headers['label']], group_key)
            tidal_volume_cm3 = validate_numeric(
                row[age_headers['tidal_volume_cm3']],
                f'tidal_volume_cm3 (row {row_number})',
            )
            breathing_freq = validate_numeric(
                row[age_headers['breathing_freq_per_min']],
                f'breathing_freq_per_min (row {row_number})',
            )
            age_groups[group_key] = AgeGroupConfig(
                key=group_key,
                label=label,
                tidal_volume_cm3=tidal_volume_cm3,
                breathing_freq=breathing_freq,
            )

        if not age_groups:
            raise ValueError('The workbook must include at least one age-group row.')

        return SimulationConfig(
            n_particles=n_particles,
            pm_sizes=pm_sizes,
            age_groups=age_groups,
            seed=seed,
        )
    finally:
        workbook.close()


def save_text_file(file_path: str | FilePath, content: str):
    """Write UTF-8 text output to disk."""
    FilePath(file_path).write_text(content, encoding='utf-8')


def run_application(
    config: SimulationConfig,
    output_dir: str | FilePath = '.',
    progress_callback: Optional[Callable[[Dict[str, object]], None]] = None,
    log_callback: Optional[Callable[[str], None]] = None,
) -> Dict[str, object]:
    """Execute the full workbook-driven simulation and write outputs."""
    config = clone_config(config)
    if not config.pm_sizes:
        raise ValueError('The simulation configuration does not contain any PM sizes.')
    if not config.age_groups:
        raise ValueError('The simulation configuration does not contain any age groups.')

    output_path = FilePath(output_dir).expanduser().resolve()
    output_path.mkdir(parents=True, exist_ok=True)

    runtime_age_groups = config.runtime_age_groups()
    age_keys = list(config.age_groups.keys())
    primary_age_key = get_primary_age_key(config)
    total_units = (
        len(age_keys) * len(config.pm_sizes) * config.n_particles
        + len(OUTPUT_FILENAMES)
        + len(age_keys) * len(AGE_GROUP_OUTPUT_PATTERNS)
    )
    completed_units = 0

    emit_log('=' * 70, log_callback)
    emit_log('  PM DEPOSITION MONTE CARLO SIMULATION', log_callback)
    emit_log('  XLSX-driven stochastic lung model with Weibel geometry', log_callback)
    emit_log(f'  Particles per simulation: {config.n_particles:,}', log_callback)
    emit_log(f'  Output directory: {output_path}', log_callback)
    emit_log('=' * 70, log_callback)
    report_progress(
        progress_callback,
        stage='initializing',
        message='Preparing simulation inputs...',
        completed_units=completed_units,
        total_units=total_units,
    )

    all_age_results = {}
    for age_key in age_keys:
        age_group = config.age_groups[age_key]
        age_display = format_age_display(age_group)
        emit_log(f'\n> Simulating age group: {age_display}', log_callback)
        emit_log(
            f'  Tidal volume: {age_group.tidal_volume_cm3:.0f} cm3, '
            f'Breathing freq: {age_group.breathing_freq}/min',
            log_callback,
        )

        age_offset = completed_units

        def on_age_progress(payload: Dict[str, object], *, current_age=age_display, current_key=age_key):
            report_progress(
                progress_callback,
                stage='simulation',
                message=(
                    f'Simulating {current_age} | PM {payload["pm_size"]} um '
                    f'({payload["particle"]}/{payload["particle_total"]})'
                ),
                completed_units=int(payload['completed_units']),
                total_units=total_units,
                age_group=current_key,
                pm_size=payload['pm_size'],
            )

        results = run_simulation(
            n_particles=config.n_particles,
            age_group=age_key,
            age_groups=runtime_age_groups,
            pm_sizes=config.pm_sizes,
            seed=config.seed,
            progress_callback=on_age_progress,
            progress_offset=age_offset,
            progress_total=total_units,
            log_callback=log_callback,
        )
        all_age_results[age_key] = results
        completed_units = age_offset + config.n_particles * len(config.pm_sizes)

    summary_sections = []
    for age_key in age_keys:
        age_group = config.age_groups[age_key]
        age_display = format_age_display(age_group)
        heading = age_key if age_display in age_key else f'{age_key} - {age_display}'
        summary_sections.append(render_results_table(all_age_results[age_key], heading, config.pm_sizes).strip())

    formula_lines = [
        '',
        "> Monte Carlo 8-Hour Exposure Formula (Paper's method):",
        '  Formula: MC = (30 x N x rho) / (TV x BF)',
    ]
    for age_key in age_keys:
        age_group = config.age_groups[age_key]
        age_display = format_age_display(age_group)
        formula_lines.append(f'  {age_display}:')
        for particle_count in [250, 1000, 2500]:
            mc_value = monte_carlo_formula(
                particle_count,
                1.0,
                age_group.tidal_volume_cm3,
                age_group.breathing_freq,
            )
            formula_lines.append(f'    N={particle_count:>5} particles -> MC factor = {mc_value:.4f}')
    full_summary_text = '\n\n'.join(summary_sections) + '\n' + '\n'.join(formula_lines) + '\n'

    summary_path = output_path / OUTPUT_FILENAMES['summary']
    save_text_file(summary_path, full_summary_text)
    emit_log(full_summary_text, log_callback)
    completed_units += 1
    report_progress(
        progress_callback,
        stage='outputs',
        message='Saved textual summary.',
        completed_units=completed_units,
        total_units=total_units,
    )

    emit_log('\n> Generating visualizations...', log_callback)
    output_files = {'summary': str(summary_path)}
    output_files['age_group_images'] = {}

    bars_path = output_path / OUTPUT_FILENAMES['bars']
    generate_deposition_bar_charts(
        all_age_results,
        config.pm_sizes,
        config.age_groups,
        config.n_particles,
        str(bars_path),
        log_callback,
    )
    output_files['bars'] = str(bars_path)
    completed_units += 1
    report_progress(
        progress_callback,
        stage='outputs',
        message='Generated age-group deposition bar charts.',
        completed_units=completed_units,
        total_units=total_units,
    )

    for age_key in age_keys:
        age_group = config.age_groups[age_key]
        age_display = format_age_display(age_group)
        age_results = all_age_results[age_key]
        age_slug = slugify_for_filename(age_key)
        age_output_files = {}

        scatter_path = output_path / AGE_GROUP_OUTPUT_PATTERNS['scatter'].format(slug=age_slug)
        generate_3d_scatter(age_results, config.pm_sizes, age_display, str(scatter_path), log_callback)
        age_output_files['scatter'] = str(scatter_path)
        completed_units += 1
        report_progress(
            progress_callback,
            stage='outputs',
            message=f'Generated 3D scatter plot for {age_display}.',
            completed_units=completed_units,
            total_units=total_units,
        )

        heatmap_path = output_path / AGE_GROUP_OUTPUT_PATTERNS['heatmap'].format(slug=age_slug)
        generate_heatmap(age_results, age_display, config.pm_sizes, str(heatmap_path), log_callback)
        age_output_files['heatmap'] = str(heatmap_path)
        completed_units += 1
        report_progress(
            progress_callback,
            stage='outputs',
            message=f'Generated heatmap for {age_display}.',
            completed_units=completed_units,
            total_units=total_units,
        )

        infographic_path = output_path / AGE_GROUP_OUTPUT_PATTERNS['infographic'].format(slug=age_slug)
        generate_lung_infographic(
            age_results,
            age_display,
            config.pm_sizes,
            pm_size=2.5 if 2.5 in config.pm_sizes else config.pm_sizes[0],
            save_path=str(infographic_path),
            log_callback=log_callback,
        )
        age_output_files['infographic'] = str(infographic_path)
        completed_units += 1
        report_progress(
            progress_callback,
            stage='outputs',
            message=f'Generated lung infographic for {age_display}.',
            completed_units=completed_units,
            total_units=total_units,
        )

        output_files['age_group_images'][age_key] = age_output_files

    primary_outputs = output_files['age_group_images'][primary_age_key]
    output_files['primary_age_group'] = primary_age_key
    output_files['scatter'] = primary_outputs['scatter']
    output_files['heatmap'] = primary_outputs['heatmap']
    output_files['infographic'] = primary_outputs['infographic']

    emit_log('\nSimulation complete!', log_callback)
    report_progress(
        progress_callback,
        stage='complete',
        message='Simulation complete.',
        completed_units=total_units,
        total_units=total_units,
    )
    return {
        'config': config,
        'results': all_age_results,
        'output_files': output_files,
        'summary_text': full_summary_text,
        'output_dir': str(output_path),
    }


class PollutionParticleApp:
    """Native desktop UI for workbook-driven simulations."""

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title('PM Deposition Simulator')
        self.root.geometry('1440x920')
        self.root.minsize(1100, 760)

        self.message_queue: queue.Queue = queue.Queue()
        self.worker_thread: Optional[threading.Thread] = None
        self.preview_images: Dict[str, ImageTk.PhotoImage] = {}

        self.input_path_var = tk.StringVar(value=str((FilePath.cwd() / DEFAULT_TEMPLATE_NAME).resolve()))
        self.output_dir_var = tk.StringVar(value=str(FilePath.cwd()))
        self.status_var = tk.StringVar(value='Select an Excel workbook and an output directory, then run the simulation.')
        self.progress_var = tk.DoubleVar(value=0.0)

        self._build_layout()
        self.root.after(100, self.process_queue)

    def _build_layout(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        controls = ttk.Frame(self.root, padding=16)
        controls.grid(row=0, column=0, sticky='ew')
        controls.columnconfigure(1, weight=1)

        ttk.Label(controls, text='XLSX Input').grid(row=0, column=0, sticky='w', padx=(0, 12), pady=(0, 8))
        ttk.Entry(controls, textvariable=self.input_path_var).grid(row=0, column=1, sticky='ew', pady=(0, 8))
        ttk.Button(controls, text='Browse...', command=self.choose_input_file).grid(row=0, column=2, padx=(12, 6), pady=(0, 8))
        ttk.Button(controls, text='Create Template', command=self.create_template).grid(row=0, column=3, pady=(0, 8))

        ttk.Label(controls, text='Output Directory').grid(row=1, column=0, sticky='w', padx=(0, 12), pady=(0, 8))
        ttk.Entry(controls, textvariable=self.output_dir_var).grid(row=1, column=1, sticky='ew', pady=(0, 8))
        ttk.Button(controls, text='Browse...', command=self.choose_output_directory).grid(row=1, column=2, padx=(12, 6), pady=(0, 8))

        self.run_button = ttk.Button(controls, text='Run Simulation', command=self.start_run)
        self.run_button.grid(row=1, column=3, pady=(0, 8))

        ttk.Label(controls, textvariable=self.status_var).grid(row=2, column=0, columnspan=4, sticky='w', pady=(4, 6))
        ttk.Progressbar(controls, variable=self.progress_var, maximum=100).grid(row=3, column=0, columnspan=4, sticky='ew')

        content = ttk.Panedwindow(self.root, orient='horizontal')
        content.grid(row=1, column=0, sticky='nsew', padx=16, pady=(0, 16))

        log_frame = ttk.Frame(content, padding=8)
        preview_frame = ttk.Frame(content, padding=8)
        content.add(log_frame, weight=1)
        content.add(preview_frame, weight=2)

        ttk.Label(log_frame, text='Run Log').pack(anchor='w')
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap='word', height=20)
        self.log_text.pack(fill='both', expand=True, pady=(8, 0))
        self.log_text.configure(state='disabled')

        ttk.Label(preview_frame, text='Output Preview').pack(anchor='w')
        self.preview_notebook = ttk.Notebook(preview_frame)
        self.preview_notebook.pack(fill='both', expand=True, pady=(8, 0))
        self.show_placeholder_preview('Run the simulation to preview the generated charts here.')

    def choose_input_file(self):
        file_path = filedialog.askopenfilename(
            title='Select simulation workbook',
            filetypes=[('Excel Workbook', '*.xlsx')],
        )
        if file_path:
            self.input_path_var.set(file_path)

    def choose_output_directory(self):
        directory = filedialog.askdirectory(title='Select output directory')
        if directory:
            self.output_dir_var.set(directory)

    def create_template(self):
        file_path = filedialog.asksaveasfilename(
            title='Create starter workbook',
            defaultextension='.xlsx',
            initialfile=DEFAULT_TEMPLATE_NAME,
            filetypes=[('Excel Workbook', '*.xlsx')],
        )
        if not file_path:
            return

        try:
            create_template_workbook(file_path)
        except Exception as exc:
            messagebox.showerror('Template Error', str(exc))
            return

        self.input_path_var.set(file_path)
        self.append_log(f'Created workbook template: {file_path}')
        self.status_var.set('Workbook template created. You can edit it and run the simulation.')

    def start_run(self):
        input_path = self.input_path_var.get().strip()
        output_dir = self.output_dir_var.get().strip()
        if not input_path:
            messagebox.showerror('Missing Input', 'Select an Excel workbook before running the simulation.')
            return
        if not output_dir:
            messagebox.showerror('Missing Output Directory', 'Select an output directory before running the simulation.')
            return
        if self.worker_thread is not None and self.worker_thread.is_alive():
            return

        self.run_button.configure(state='disabled')
        self.progress_var.set(0.0)
        self.status_var.set('Loading workbook...')
        self.clear_previews()
        self.append_log(f'Loading workbook: {input_path}')
        self.append_log(f'Output directory: {output_dir}')

        self.worker_thread = threading.Thread(
            target=self._run_worker,
            args=(input_path, output_dir),
            daemon=True,
        )
        self.worker_thread.start()

    def _run_worker(self, input_path: str, output_dir: str):
        try:
            config = load_config_from_workbook(input_path)
            self.message_queue.put({'type': 'log', 'text': f'Loaded workbook: {input_path}'})
            result = run_application(
                config,
                output_dir,
                progress_callback=lambda payload: self.message_queue.put({'type': 'progress', 'payload': payload}),
                log_callback=lambda text: self.message_queue.put({'type': 'log', 'text': text}),
            )
            self.message_queue.put({'type': 'done', 'result': result})
        except Exception as exc:
            self.message_queue.put({
                'type': 'error',
                'message': str(exc),
                'details': traceback.format_exc(),
            })

    def process_queue(self):
        while True:
            try:
                item = self.message_queue.get_nowait()
            except queue.Empty:
                break

            item_type = item['type']
            if item_type == 'log':
                self.append_log(item['text'])
            elif item_type == 'progress':
                payload = item['payload']
                self.status_var.set(payload.get('message', 'Running simulation...'))
                self.progress_var.set(float(payload.get('progress', 0.0)) * 100)
            elif item_type == 'done':
                self.run_button.configure(state='normal')
                self.progress_var.set(100.0)
                self.status_var.set(f"Completed. Outputs saved to {item['result']['output_dir']}")
                self.append_log(f"Outputs saved to {item['result']['output_dir']}")
                self.show_output_previews(item['result'])
            elif item_type == 'error':
                self.run_button.configure(state='normal')
                self.status_var.set('Simulation failed. Check the log for details.')
                self.append_log(item['details'])
                messagebox.showerror('Simulation Error', item['message'])

        self.root.after(100, self.process_queue)

    def append_log(self, text: str):
        self.log_text.configure(state='normal')
        self.log_text.insert('end', text.rstrip() + '\n')
        self.log_text.see('end')
        self.log_text.configure(state='disabled')

    def clear_previews(self):
        self.preview_images.clear()
        for tab_id in self.preview_notebook.tabs():
            self.preview_notebook.forget(tab_id)
        self.show_placeholder_preview('Run the simulation to preview the generated charts here.')

    def show_placeholder_preview(self, message: str):
        placeholder = ttk.Frame(self.preview_notebook)
        ttk.Label(placeholder, text=message, anchor='center', justify='center').pack(fill='both', expand=True, padx=16, pady=16)
        self.preview_notebook.add(placeholder, text='Preview')

    def add_image_preview_tab(self, tab_title: str, image_path: FilePath, image_key: str):
        """Add one rendered output image to the preview notebook."""
        if not image_path.exists():
            return

        tab = ttk.Frame(self.preview_notebook)
        with Image.open(image_path) as image:
            pil_image = image.copy()
        pil_image.thumbnail((980, 680))
        photo = ImageTk.PhotoImage(pil_image)
        self.preview_images[image_key] = photo

        ttk.Label(tab, image=photo).pack(fill='both', expand=True, padx=12, pady=(12, 6))
        ttk.Label(tab, text=str(image_path)).pack(anchor='w', padx=12, pady=(0, 12))
        self.preview_notebook.add(tab, text=tab_title)

    def show_output_previews(self, result: Dict[str, object]):
        for tab_id in self.preview_notebook.tabs():
            self.preview_notebook.forget(tab_id)
        self.preview_images.clear()

        summary_tab = ttk.Frame(self.preview_notebook)
        summary_text = scrolledtext.ScrolledText(summary_tab, wrap='word', height=18)
        summary_text.pack(fill='both', expand=True)
        summary_text.insert('1.0', result['summary_text'])
        summary_text.configure(state='disabled')
        self.preview_notebook.add(summary_tab, text='Summary')

        self.add_image_preview_tab('Bar Charts', FilePath(result['output_files']['bars']), 'bars')

        age_group_images = result['output_files'].get('age_group_images', {})
        for age_key, age_group in result['config'].age_groups.items():
            if age_key not in age_group_images:
                continue

            age_display = format_age_display(age_group)
            age_slug = slugify_for_filename(age_key)
            for image_key, image_label in [
                ('scatter', '3D Scatter'),
                ('heatmap', 'Heatmap'),
                ('infographic', 'Infographic'),
            ]:
                image_path = age_group_images[age_key].get(image_key)
                if not image_path:
                    continue
                self.add_image_preview_tab(
                    f'{age_display} | {image_label}',
                    FilePath(image_path),
                    f'{image_key}_{age_slug}',
                )


def launch_ui():
    """Start the desktop application."""
    root = tk.Tk()
    PollutionParticleApp(root)
    root.mainloop()


def parse_args(argv: Optional[List[str]] = None):
    """Parse CLI arguments for workbook, template, and UI modes."""
    parser = argparse.ArgumentParser(description='PM deposition simulation with XLSX input and desktop UI.')
    parser.add_argument('--ui', action='store_true', help='Launch the desktop UI.')
    parser.add_argument('--xlsx', help='Path to the simulation workbook (.xlsx).')
    parser.add_argument('--output-dir', default='.', help='Directory for generated outputs.')
    parser.add_argument(
        '--generate-template',
        nargs='?',
        const=DEFAULT_TEMPLATE_NAME,
        help='Write a starter workbook and optionally choose the destination path.',
    )
    parser.add_argument(
        '--run-defaults',
        action='store_true',
        help='Run the simulation with the built-in default data instead of loading a workbook.',
    )
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None):
    """Entry point for CLI and desktop usage."""
    args = parse_args(argv)

    if args.generate_template is not None:
        template_path = FilePath(args.generate_template)
        create_template_workbook(template_path)
        emit_log(f'Created workbook template: {template_path.resolve()}')
        if not args.ui and not args.xlsx and not args.run_defaults:
            return {'template': str(template_path.resolve())}

    if args.ui or (not args.xlsx and not args.run_defaults):
        launch_ui()
        return None

    if args.run_defaults:
        config = get_default_config()
    else:
        config = load_config_from_workbook(args.xlsx)
    return run_application(config, args.output_dir)


if __name__ == '__main__':
    main()